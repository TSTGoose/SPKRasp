import hashlib
from pathlib import Path
from pprint import pprint
from openpyxl import load_workbook


class Week:
    weektypes: list = ["Нечётная", "Чётная"]
    weekdays: list = [
        "Понедельник",
        "Вторник",
        "Среда",
        "Четверг",
        "Пятница",
        "Суббота"
    ]


class Lesson:
    pass


class Parser(object):

    def __init__(self, path: Path, filename: str = "current_rasp.xlsx"):
        self.document = load_workbook(f"{path}/{filename}", read_only=True)
        self.sheet = self.document["Основное расписание"]
        self.weektypes = Week.weektypes

    def start_parsing(self):
        print("Начало парсинга.")
        groups: list = self._parsing_groups() # Список групп
        rasp_tamplate: dict = self._create_tamplate(groups=groups) # Шаблон расписания
        for group, group_col in zip(groups, range(6, 235, 5)):
            for weekday, weekday_row in zip(Week.weekdays, range(8, 80, 12)):
                print(f"Парсинг расписания за {weekday} для группы {group}")
                rasp_tamplate[group]["Нечётная"][weekday] = self._parsing_lessons(start_row=weekday_row, start_col=group_col)
                rasp_tamplate[group]["Чётная"][weekday] = self._parsing_lessons(start_row=weekday_row+75, start_col=group_col)
        return rasp_tamplate

    # Создаёт шаблон расписания
    def _create_tamplate(self, groups: list) -> dict:
        print("Создание шаблона.")
        tamplate: dict = {group: {} for group in groups}

        # Нечётная неделя начинается с 8 строки и заканчивается на 80.
        # Чётная неделя начинается с 83
        for weektype, row in zip(self.weektypes, range(8, 84, 75)):
            print(f"Создаётся шаблон. {weektype} неделя")
            for col, group in zip(range(5, 235, 5), groups):
                tamplate_draft = {weektype: {weekday: dict for weekday in Week.weekdays}}
                tamplate[group].update(tamplate_draft)
        return tamplate

    # Создаёт список групп.
    def _parsing_groups(self) -> list:
        print("Парсинг групп")
        groups = []
        for row in self.sheet.iter_rows(min_row=6, max_row=6, min_col=5, max_col=234, values_only=True):
            groups = [cell for cell in row if cell is not None]
        return groups

    # Возвращает словарь с парами на день недели
    def _parsing_lessons(self, start_row: int = 8, start_col: int = 6) -> dict:
        end_col = start_col + 3
        end_row = start_row + 12
        lessons = {}

        for row, lesson_index in zip(range(start_row, end_row, 2), range(1, 7)):
            lesson_in_doc = [row for row in self.sheet.iter_rows(min_row=row, max_row=row + 1,
                                                                 min_col=start_col, max_col=end_col,
                                                                 values_only=True)]
            check_subgroup_descipline: bool = lesson_in_doc[0][0]
            check_subgroup_kab: bool = lesson_in_doc[0][1]

            # Если дисциплина для первой подгруппы есть, но не стоит кабинет - значит пара не поделена на подгруппы
            if check_subgroup_descipline and not check_subgroup_kab:
                lessons[lesson_index] = {
                    "descipline": lesson_in_doc[0][0],
                    "kab": lesson_in_doc[0][3],
                    "teacher": lesson_in_doc[1][0]
                }

            # Если диспилина для первой и второй подгруппы не стоит, значит пары нет
            elif lesson_in_doc[0][0] is None and lesson_in_doc[0][2] is None:
                lessons[lesson_index] = {"Пары нет"}

            # Иначе, пара поделена на подгруппы
            else:
                lessons[lesson_index] = {
                    "first_subgroup": {
                        "descipline": lesson_in_doc[0][0],
                        "kab": lesson_in_doc[0][1],
                        "teacher": lesson_in_doc[1][0]
                    },
                    "second_subgroup": {
                        "descipline": lesson_in_doc[0][2],
                        "kab": lesson_in_doc[0][3],
                        "teacher": lesson_in_doc[1][2]
                    }
                }
        return lessons

    # Сравнение расписания
    def compare_rasp(self, new_rasp):
        new_rasp = load_workbook(new_rasp, read_only=True, data_only=True)
        current_rasp = self.document
        new_rasp_data = [doc_value for doc_value in new_rasp["Основное расписание"].iter_rows(values_only=True)]
        current_rasp_data = [doc_value for doc_value in current_rasp["Основное расписание"].iter_rows(values_only=True)]
        is_changed = new_rasp_data != current_rasp_data
        new_rasp.close()
        current_rasp.close()
        return is_changed
